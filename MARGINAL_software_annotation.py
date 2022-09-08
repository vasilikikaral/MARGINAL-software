import openpyxl 
from openpyxl.utils import get_column_letter, column_index_from_string 
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import sys
import os
import scipy
from scipy import stats
from scipy.stats import fisher_exact



user_input = input("Enter the path of your file (with double \): ")
input_file = input("Enter the file name with file extension: ")
path_filtered = "{}\\{}".format(user_input,input_file)
assert os.path.exists(user_input), "I did not find the file at, "+str(user_input)


initial_workbook = path_filtered

df_initial = pd.read_excel(path_filtered)

df_initial_new = df_initial.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(keep='first')

df_initial_new.to_excel(initial_workbook, index=False)



#for i in range(m_row, 1, -1):       
     #cell_obj_feature = sheet_obj.cell(row = i, column = 9)
     #if (cell_obj_feature.value != "NM_000059.3") and (cell_obj_feature.value != "NM_007294.4"):
        #sheet_obj.delete_rows(i, 1)
        #print(cell_obj_feature.value)
#wb_obj.save(path_filtered)



df_initial = pd.read_excel(path_filtered)

#df_initial = df_initial.sort_values('#Uploaded_variation',ascending=True)
df_initial = df_initial.sort_values('Location',ascending=True)

df_initial.to_excel(initial_workbook, index=False)


wb_obj = openpyxl.load_workbook(initial_workbook) 
   
sheet_obj = wb_obj.active 

m_row = sheet_obj.max_row
max_col = sheet_obj.max_column





#PP3_BP4

val8 = input("Enter the letter of column MetaSVM_pred: ") 
print(val8)
index_MetaSVM_pred = column_index_from_string(val8)


val9 = input("Enter the letter of column Condel: ") 
print(val9)
index_Condel = column_index_from_string(val9)


#val10 = input("Enter the letter of column GERP++_NR: ") 
#print(val10)
#index_GERP_NR = column_index_from_string(val10)


val11 = input("Enter the letter of column GERP++_RS: ") 
print(val11)
index_GERP_RS = column_index_from_string(val11)


val12 = input("Enter the letter of column ada_score: ") 
print(val12)
index_ada_score = column_index_from_string(val12)


val13 = input("Enter the letter of column rf_score: ") 
print(val13)
index_rf_score = column_index_from_string(val13)


result_pp3 = sheet_obj.cell(row= 1, column= max_col + 1, value='PP3')

result_bp4 = sheet_obj.cell(row= 1, column= max_col + 2, value='BP4')



for i in range(2, m_row + 1):

       cell_obj_MetaSVM_pred = sheet_obj.cell(row = i, column = index_MetaSVM_pred)
       #print(cell_obj_MetaSVM_pred.value)

       #cell_obj_Condel = sheet_obj.cell(row = i, column = index_Condel)
       #print(cell_obj_Condel.value)

       #cell_obj_GERP_NR = sheet_obj.cell(row = i, column = index_GERP_NR)
       #print(cell_obj_GERP_NR.value)

       cell_obj_GERP_RS = sheet_obj.cell(row = i, column = index_GERP_RS)
       #print(cell_obj_GERP_RS.value)

       cell_obj_ada_score = sheet_obj.cell(row = i, column = index_ada_score)
       #print(cell_obj_ada_score.value)

       cell_obj_rf_score = sheet_obj.cell(row = i, column = index_rf_score)
       #print(cell_obj_rf_score.value)



       if (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value > 2) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value > 2):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value > 2) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value > 2) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value > 2):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value > 2):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value > 2) and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-"):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('D' in cell_obj_MetaSVM_pred.value):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_ada_score.value > 0.6) and (cell_obj_rf_score.value > 0.6):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value > 2):
          sheet_obj.cell(row = i, column = max_col + 1).value = 1
          sheet_obj.cell(row = i, column = max_col + 2).value = 0





       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value < 2) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value < 2):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value < 2) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value < 2) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value < 2):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value) and (cell_obj_GERP_RS.value < 2):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value < 2) and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-"):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value != "-") and ('T' in cell_obj_MetaSVM_pred.value):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_MetaSVM_pred.value == "-") and (cell_obj_GERP_RS.value < 2):
          sheet_obj.cell(row = i, column = max_col + 2).value = 1
          sheet_obj.cell(row = i, column = max_col + 1).value = 0


       elif (cell_obj_GERP_RS.value == "-") and (cell_obj_ada_score.value == "-") and (cell_obj_rf_score.value == "-") and (cell_obj_MetaSVM_pred.value == "-"):

          sheet_obj.cell(row = i, column = max_col + 1).value = 0
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


       else:
          sheet_obj.cell(row = i, column = max_col + 1).value = 0
          sheet_obj.cell(row = i, column = max_col + 2).value = 0


wb_obj.save(initial_workbook)











#PVS1


#result_pvs1 = sheet_obj.cell(row= 1, column= max_col + 3, value='PVS1')

df_initial = pd.read_excel(initial_workbook)

df_initial[['Location_start','Location_end']] = df_initial['Location'].str.split("-",expand=True)

df_initial.to_excel(initial_workbook, index=False)


df_initial = pd.read_excel(initial_workbook)

#all except the last exon of BRCA2

df_initial['PVS1'] = np.where(((df_initial["Consequence"] == "splice_donor_variant") | (df_initial["Consequence"] == "stop_gained") | (df_initial["Consequence"] == "stop_lost") | (df_initial["Consequence"] == "start_lost") | (df_initial["Consequence"] == "frameshift_variant") | (df_initial["Consequence"] == "splice_acceptor_variant") | (df_initial["Consequence"] == "stop_gained,frameshift_variant") | (df_initial["Consequence"] == "frameshift_variant,stop_lost")) & ((df_initial["Location_end"].astype(int) < 32973746) | (df_initial["Location_end"].astype(int) > 32973805)), 1, 0)



df_initial.to_excel("Enter your path\\VEP_variants_Final.xlsx", index=False)





#BA1


wb_obj = openpyxl.load_workbook(initial_workbook) 
   
sheet_obj = wb_obj.active 

m_row = sheet_obj.max_row
max_col = sheet_obj.max_column



val6 = input("Enter the letter of column ExAC_nonTCGA_AF: ") 
print(val6)
index_ExAC_nonTCGA_AF = column_index_from_string(val6)

val7 = input("Enter the letter of column gnomAD_exomes_controls_AF: ") 
print(val7)
index_gnomAD_exomes_controls_AF = column_index_from_string(val7)

#val17 = input("Enter the letter of column 1000Gp3_EUR_AF: ") 
#print(val17)
#index_1000Gp3_EUR_AF = column_index_from_string(val17)

result_ba1 = sheet_obj.cell(row= 1, column= max_col + 1, value='BA1')


for i in range(2, m_row + 1):

       cell_obj_ExAC_nonTCGA_AF = sheet_obj.cell(row = i, column = index_ExAC_nonTCGA_AF)
       print(cell_obj_ExAC_nonTCGA_AF.value)

       cell_obj_gnomAD_exomes_controls_AF = sheet_obj.cell(row = i, column = index_gnomAD_exomes_controls_AF)
       print(cell_obj_gnomAD_exomes_controls_AF.value)
       
       if (cell_obj_ExAC_nonTCGA_AF.value != "-") and (cell_obj_gnomAD_exomes_controls_AF.value != "-"):

           if (cell_obj_ExAC_nonTCGA_AF.value > 0.05) or (cell_obj_gnomAD_exomes_controls_AF.value > 0.05):
               #print(sheet_obj.cell(row = i, column = 1).value, 'BA1: 1')
               sheet_obj.cell(row = i, column = max_col + 1).value = 1

           else: 
               #print(sheet_obj.cell(row = i, column = 1).value, 'BA1: 0')
               sheet_obj.cell(row = i, column = max_col + 1).value = 0

       elif (cell_obj_ExAC_nonTCGA_AF.value != "-") and (cell_obj_ExAC_nonTCGA_AF.value > 0.05) and (cell_obj_gnomAD_exomes_controls_AF.value == "-"): 
          sheet_obj.cell(row = i, column = max_col + 1).value = 1


       elif (cell_obj_gnomAD_exomes_controls_AF.value != "-") and (cell_obj_gnomAD_exomes_controls_AF.value > 0.05) and (cell_obj_ExAC_nonTCGA_AF.value == "-"): 
          sheet_obj.cell(row = i, column = max_col + 1).value = 1


       else: 
               #print(sheet_obj.cell(row = i, column = 1).value, 'BA1: 0')
               sheet_obj.cell(row = i, column = max_col + 1).value = 0


wb_obj.save(initial_workbook)






#BS1

result_bs1 = sheet_obj.cell(row= 1, column= max_col + 2, value='BS1')


for i in range(2, m_row + 1):

       cell_obj_ExAC_nonTCGA_AF = sheet_obj.cell(row = i, column = index_ExAC_nonTCGA_AF)
       print(cell_obj_ExAC_nonTCGA_AF.value)

       cell_obj_gnomAD_exomes_controls_AF = sheet_obj.cell(row = i, column = index_gnomAD_exomes_controls_AF)
       print(cell_obj_gnomAD_exomes_controls_AF.value)

       if ((cell_obj_ExAC_nonTCGA_AF.value != "-") and (cell_obj_gnomAD_exomes_controls_AF.value != "-")):

           if ((0.01 < cell_obj_ExAC_nonTCGA_AF.value) or (0.01 < cell_obj_gnomAD_exomes_controls_AF.value)):
              #print(sheet_obj.cell(row = i, column = 1).value, 'BS1: 1')
              sheet_obj.cell(row = i, column = max_col + 2).value = 1

           else: 
              #print(sheet_obj.cell(row = i, column = 1).value, 'BS1: 0')
              sheet_obj.cell(row = i, column = max_col + 2).value = 0

       elif ((cell_obj_ExAC_nonTCGA_AF.value != "-") and (0.01 < cell_obj_ExAC_nonTCGA_AF.value) and (cell_obj_gnomAD_exomes_controls_AF.value == "-")):
           sheet_obj.cell(row = i, column = max_col + 2).value = 1 


       elif ((cell_obj_gnomAD_exomes_controls_AF.value != "-") and (0.01 < cell_obj_gnomAD_exomes_controls_AF.value) and (cell_obj_ExAC_nonTCGA_AF.value == "-")):
           sheet_obj.cell(row = i, column = max_col + 2).value = 1


       else: 
              #print(sheet_obj.cell(row = i, column = 1).value, 'BS1: 0')
              sheet_obj.cell(row = i, column = max_col + 2).value = 0


wb_obj.save(initial_workbook)






#PM2

result_pm2 = sheet_obj.cell(row= 1, column= max_col + 3, value='PM2')


for i in range(2, m_row + 1):

       cell_obj_ExAC_nonTCGA_AF = sheet_obj.cell(row = i, column = index_ExAC_nonTCGA_AF)
       print(cell_obj_ExAC_nonTCGA_AF.value)

       cell_obj_gnomAD_exomes_controls_AF = sheet_obj.cell(row = i, column = index_gnomAD_exomes_controls_AF)
       print(cell_obj_gnomAD_exomes_controls_AF.value)

       if ((cell_obj_ExAC_nonTCGA_AF.value == 0) or (cell_obj_ExAC_nonTCGA_AF.value == "-")):
           if ((cell_obj_gnomAD_exomes_controls_AF.value == 0) or (cell_obj_gnomAD_exomes_controls_AF.value == "-")):
              #print(sheet_obj.cell(row = i, column = 1).value, 'PM2: 1')
              sheet_obj.cell(row = i, column = max_col + 3).value = 1

           else: 
              #print(sheet_obj.cell(row = i, column = 1).value, 'PM2: 0')
              sheet_obj.cell(row = i, column = max_col + 3).value = 0

       else: 
              #print(sheet_obj.cell(row = i, column = 1).value, 'PM2: 0')
              sheet_obj.cell(row = i, column = max_col + 3).value = 0


wb_obj.save(initial_workbook)






#BP7


result_bp7 = sheet_obj.cell(row= 1, column= max_col + 4, value='BP7')


for i in range(2, m_row + 1):

       cell_obj_Consequence = sheet_obj.cell(row = i, column = 7)

       #cell_obj_GERP_NR = sheet_obj.cell(row = i, column = index_GERP_NR)

       cell_obj_GERP_RS = sheet_obj.cell(row = i, column = index_GERP_RS)

       cell_obj_ada_score = sheet_obj.cell(row = i, column = index_ada_score)

       cell_obj_rf_score = sheet_obj.cell(row = i, column = index_rf_score)

       if (('synonymous' in cell_obj_Consequence.value) and (cell_obj_GERP_RS.value != "-") and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6) and (cell_obj_GERP_RS.value < 2)):
          sheet_obj.cell(row = i, column = max_col + 4).value = 1



       elif (('synonymous' in cell_obj_Consequence.value) and (cell_obj_GERP_RS.value != "-") and ((cell_obj_ada_score.value == "-") or (cell_obj_rf_score.value == "-")) and (cell_obj_GERP_RS.value < 2)):
          sheet_obj.cell(row = i, column = max_col + 4).value = 1


       elif (('synonymous' in cell_obj_Consequence.value) and ((cell_obj_GERP_RS.value == "-")) and (cell_obj_ada_score.value != "-") and (cell_obj_rf_score.value != "-") and (cell_obj_ada_score.value < 0.6) and (cell_obj_rf_score.value < 0.6)):
          sheet_obj.cell(row = i, column = max_col + 4).value = 1


       else: 
          sheet_obj.cell(row = i, column = max_col + 4).value = 0



wb_obj.save(initial_workbook)






#PP2_BP1

result_pp2 = sheet_obj.cell(row= 1, column= max_col + 5, value='PP2')

result_bp1 = sheet_obj.cell(row= 1, column= max_col + 6, value='BP1')


for i in range(2, m_row + 1):

       #print(sheet_obj.cell(row = i, column = 1).value, 'PP2: 0')
       sheet_obj.cell(row = i, column = max_col + 5).value = 0

       #print(sheet_obj.cell(row = i, column = 1).value, 'BP1: 0')
       sheet_obj.cell(row = i, column = max_col + 6).value = 0



wb_obj.save(initial_workbook)






#PM1

result_pm1 = sheet_obj.cell(row= 1, column= max_col + 7, value='PM1')

for i in range(2, m_row + 1):

       #print(sheet_obj.cell(row = i, column = 1).value, 'PM1: 0')
       sheet_obj.cell(row = i, column = max_col + 7).value = 0



wb_obj.save(initial_workbook)











#Load the annotation file from clinvar

print('Save "clinvar_roi_annotation.xlsx" and vcf file of your data (only the first 5 columns with the headers) as an excel file to your path!')

initial_workbook = 'Enter your path\\VEP_variants_Final.xlsx'



##initial_vcf_workbook = input("Enter the path of the vcf file: ")
initial_vcf_workbook = 'Enter your path\\VEP_Total_vcf.xlsx'



vcf_output_workbook = 'Enter your path\\vcf_output.xlsx'


clinvar_annotation_file = 'Enter your path\\clinvar_roi_annotation.xlsx'

output_workbook = 'Enter your path\\output_annotation.xlsx'

df_initial = pd.read_excel(initial_workbook)

#df_initial = df_initial.sort_values('#Uploaded_variation',ascending=True)
df_initial = df_initial.sort_values('Location',ascending=True)

df_initial.to_excel("Enter your path\\VEP_variants_Final.xlsx", index=False)




df_vcf = pd.read_excel(initial_vcf_workbook)


df_clinvar = pd.read_excel(clinvar_annotation_file)

df_vcf.rename(columns={'ID':'#Uploaded_variation'}, inplace=True)

df_vcf = df_vcf.sort_values('POS',ascending=True)
#df_vcf = df_vcf.sort_values('ID',ascending=True)

df_vcf.to_excel(initial_vcf_workbook, index=False)


df_3 = pd.merge(df_initial, df_vcf[['#Uploaded_variation', 'CHROM', 'POS', 'REF', 'ALT']], on='#Uploaded_variation', how='left')
#print(df_3)
df_3.to_excel(vcf_output_workbook, index=False)

df_vcf_output = pd.read_excel(vcf_output_workbook)

df_4 = pd.merge(df_vcf_output, df_clinvar[["CHROM", "POS", "REF", "ALT", "CLNDN", "CLNHGVS", "CLNREVSTAT", "CLNSIG", "CLNVI"]], on=['CHROM', 'POS', 'REF', 'ALT'], how='left')


#print(df_4)

df_4.to_excel(output_workbook, index=False)



wb_obj = openpyxl.load_workbook(output_workbook) 
   
sheet_obj = wb_obj.active 

m_row = sheet_obj.max_row
max_col = sheet_obj.max_column


#PP5_BP6

val14 = input("Enter the letter of column 'CLNSIG': ") 
print(val14)
index_clinvar_clnsig = column_index_from_string(val14)


val15 = input("Enter the letter of column 'CLNREVSTAT': ") 
print(val15)
index_clinvar_review = column_index_from_string(val15)



result_pp5 = sheet_obj.cell(row= 1, column= max_col + 1, value='PP5')

result_bp6 = sheet_obj.cell(row= 1, column= max_col + 2, value='BP6')




for i in range(2, m_row + 1):

       cell_obj_clinvar_clnsig = sheet_obj.cell(row = i, column = index_clinvar_clnsig)
       print(cell_obj_clinvar_clnsig.value)

       cell_obj_clinvar_review = sheet_obj.cell(row = i, column = index_clinvar_review)
       print(cell_obj_clinvar_review.value)

       #cell_obj_clinvar_OMIM_id = sheet_obj.cell(row = i, column = index_clinvar_OMIM_id)
       #print(cell_obj_clinvar_OMIM_id.value)


       if (cell_obj_clinvar_clnsig.value != None):

          if ('Pathogenic' in cell_obj_clinvar_clnsig.value or 'Likely_pathogenic' in cell_obj_clinvar_clnsig.value) and (cell_obj_clinvar_review.value != "criteria_provided,_single_submitter") and (cell_obj_clinvar_review.value != "criteria_provided,_conflicting_interpretations"):
             sheet_obj.cell(row = i, column = max_col + 1).value = 1
             sheet_obj.cell(row = i, column = max_col + 2).value = 0


          elif ('Benign' in cell_obj_clinvar_clnsig.value or 'Likely_benign' in cell_obj_clinvar_clnsig.value) and (cell_obj_clinvar_review.value != "criteria_provided,_single_submitter") and (cell_obj_clinvar_review.value != "criteria_provided,_conflicting_interpretations"):
             sheet_obj.cell(row = i, column = max_col + 2).value = 1
             sheet_obj.cell(row = i, column = max_col + 1).value = 0


          elif ('Uncertain' in cell_obj_clinvar_clnsig.value or 'Conflicting' in cell_obj_clinvar_clnsig.value) or (cell_obj_clinvar_clnsig.value == "?") or (cell_obj_clinvar_clnsig.value == "not_provided") or (cell_obj_clinvar_review.value == "criteria_provided,_single_submitter"):
             sheet_obj.cell(row = i, column = max_col + 1).value = 0
             sheet_obj.cell(row = i, column = max_col + 2).value = 0

          else:
             sheet_obj.cell(row = i, column = max_col + 1).value = 0
             sheet_obj.cell(row = i, column = max_col + 2).value = 0

       else:
            sheet_obj.cell(row = i, column = max_col + 1).value = 0
            sheet_obj.cell(row = i, column = max_col + 2).value = 0


wb_obj.save(output_workbook)






#PS4

#Load the Fisher_exact_test_FINAL_patients_AC.xlsx file!

Fish_ex_t_FINAL_patients_workbook = 'Enter your path\\Fisher_exact_test_FINAL_patients_AC.xlsx'

df_initial = pd.read_excel(initial_workbook)

df_initial[['chromosome','VariantOnGenome/DNA']] = df_initial['#Uploaded_variation'].str.split(":",expand=True)

df_initial.to_excel("Enter your path\\VEP_variants_Final_new.xlsx", index=False)

df_initial_new = pd.read_excel("Enter your path\\VEP_variants_Final_new.xlsx")

df_Fish_ex_t_FINAL_patients = pd.read_excel(Fish_ex_t_FINAL_patients_workbook)

df_7 = pd.merge(df_Fish_ex_t_FINAL_patients[['VariantOnGenome/DNA', 'Case_Allele_Count', 'case_total_all']], df_initial_new[['VariantOnGenome/DNA', '#Uploaded_variation', 'gnomAD_exomes_controls_AC', 'gnomAD_exomes_controls_AN']], on='VariantOnGenome/DNA', how='outer')

Fisher_exact_test_workbook = 'Enter your path\\Fisher_exact_test_data.xlsx'

df_7.to_excel(Fisher_exact_test_workbook, index=False)

df_Fisher_data = pd.read_excel(Fisher_exact_test_workbook)


df_Fisher_data.replace(to_replace ="-", value =0, inplace = True)

df_Fisher_data.fillna(" ",inplace=True)


df_Fisher_data.loc[df_Fisher_data['gnomAD_exomes_controls_AC'] == " ", 'PS4'] = 1

df_Fisher_data.loc[df_Fisher_data['Case_Allele_Count'] == " ", 'PS4'] = 0

df_Fisher_data.to_excel(Fisher_exact_test_workbook, index=False)


df_Fisher_data['FishersExact'] = df_Fisher_data[(df_Fisher_data['PS4'].isnull())].apply(lambda r: scipy.stats.fisher_exact([[r.Case_Allele_Count, r.gnomAD_exomes_controls_AC], [r.case_total_all, r.gnomAD_exomes_controls_AN]]), axis=1)

df_Fisher_data.to_excel(Fisher_exact_test_workbook, index=False)

df_Fisher_data[['FishersExact_oddsratio','FishersExact_pvalue']] = pd.DataFrame(df_Fisher_data[(df_Fisher_data['PS4'].isnull())].FishersExact.tolist(), index= df_Fisher_data[(df_Fisher_data['PS4'].isnull())].index)

##df_Fisher_data.to_excel(Fisher_exact_test_workbook, index=False)


df_Fisher_data.loc[(df_Fisher_data['FishersExact'].notnull()) & (df_Fisher_data["FishersExact_oddsratio"] > 2) & (df_Fisher_data["FishersExact_pvalue"] < 0.05), 'PS4'] = 1

df_Fisher_data.loc[(df_Fisher_data['FishersExact'].notnull()) & ((df_Fisher_data["FishersExact_oddsratio"] < 2) | (df_Fisher_data["FishersExact_oddsratio"].isnull())), 'PS4'] = 0

df_Fisher_data.loc[(df_Fisher_data['FishersExact'].notnull()) & (df_Fisher_data["FishersExact_oddsratio"] > 2) & (df_Fisher_data["FishersExact_pvalue"] > 0.05), 'PS4'] = 0


df_Fisher_data.to_excel(Fisher_exact_test_workbook, index=False)


df_Fisher_data = pd.read_excel(Fisher_exact_test_workbook)

df_Fisher_data_new = df_Fisher_data.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(keep='first')

df_Fisher_data_new.to_excel(Fisher_exact_test_workbook, index=False)






#PM4_BP3

print('Save the output file of RepeatMasker for your data as an excel file to your path!')

output_repeatmasker_path = input("Enter the full path of the RepeatMasker file with the file name: ")


df_output_workbook = pd.read_excel(output_workbook)
df_output_repeatmasker = pd.read_excel(output_repeatmasker_path)

genstart_list = df_output_repeatmasker['genoStart'].tolist()
genend_list = df_output_repeatmasker['genoEnd'].tolist()

# Sorting the lists
genstart_list.sort()
genend_list.sort()


def createList(r1, r2):
    return np.arange(r1, r2+1, 1)


result = []

for i in range(len(genstart_list)):
  
    r1, r2 = genstart_list[i], genend_list[i]


    for element in createList(r1, r2):
        result.append(element)


df_output_workbook['Repeat'] = df_output_workbook.POS.isin(result).astype(int)

df_output_workbook.to_excel(output_workbook, index=False)


df_output_workbook['PM4'] = np.where(((df_output_workbook["Consequence"] == "inframe_insertion") | (df_output_workbook["Consequence"] == "inframe_deletion") | (df_output_workbook["Consequence"] == "stop_lost") | (df_output_workbook["Consequence"] == "frameshift_variant,stop_lost")) & (df_output_workbook["Repeat"] == 0), 1, 0)


#df_output_workbook['PM4'] = np.where((df_output_workbook["Consequence"] == "stop_lost"), 1, 0)


df_output_workbook['BP3'] = np.where(((df_output_workbook["Consequence"] == "inframe_insertion") | (df_output_workbook["Consequence"] == "inframe_deletion")) & (df_output_workbook["Repeat"] == 1), 1, 0)

df_output_workbook.to_excel(output_workbook, index=False)

df_output_workbook = pd.read_excel(output_workbook)

df_output_workbook_new = df_output_workbook.sort_values('Location',ascending=True).drop_duplicates(keep='first')

df_output_workbook_new.to_excel(output_workbook, index=False)






#PS1_PM5

#Save 'output_vep_clinvar_missense_score_filtered.xlsx' file to your path!
print('Save "output_vep_clinvar_missense_score_filtered.xlsx" file to your path!')

#output_vep_clinvar_score_workbook = input("Enter the path of the file: ")

output_vep_clinvar_score_workbook = 'Enter your path\\output_vep_clinvar_missense_score_filtered.xlsx'


output_vep_final_pathscore_workbook = 'Enter your path\\output_vep_final_missense_path.xlsx'



df_output_vep_clinvar_score = pd.read_excel(output_vep_clinvar_score_workbook)

df_5 = pd.merge(df_initial, df_output_vep_clinvar_score[["Consequence", "CDS_position", "Protein_position", "Amino_acids", "GIVEN_REF", "CLNREVSTAT", "CLNSIG", "CLNVI", "pathogenicity"]], on=['Consequence', 'CDS_position', 'Protein_position', 'Amino_acids', 'GIVEN_REF'], how='left')

#df_5 = pd.merge(df_initial, df_output_vep_clinvar_score[["Consequence", "CDS_position", "Protein_position", "Amino_acids", "HGVSc_ref", "CLNREVSTAT", "CLNSIG", "CLNVI", "pathogenicity"]], on=['Consequence', 'CDS_position', 'Protein_position', 'Amino_acids', 'HGVSc_ref'], how='left')


df_5.to_excel(output_vep_final_pathscore_workbook, index=False)

print('The "output_vep_final_missense_path.xlsx" file was saved to your path!')

wb_obj1 = openpyxl.load_workbook(output_vep_final_pathscore_workbook)
   
sheet_obj1 = wb_obj1.active 

m_row1 = sheet_obj1.max_row
max_col1 = sheet_obj1.max_column


val16 = input("Enter the letter of column 'CLNSIG' in 'output_vep_final_missense_path.xlsx' file: ") 
print(val16)
index_clinvar_clnsig = column_index_from_string(val16)


val17 = input("Enter the letter of column 'CLNREVSTAT' in 'output_vep_final_missense_path.xlsx' file: ") 
print(val17)
index_clinvar_review = column_index_from_string(val17)


val18 = input("Enter the letter of column 'pathogenicity' in 'output_vep_final_missense_path.xlsx' file: ") 
print(val18)
index_clinvar_pathogenicity = column_index_from_string(val18)

result_ps1 = sheet_obj1.cell(row= 1, column= max_col1 + 1, value='PS1')


for i in range(2, m_row1 + 1):

       cell_obj1_clinvar_clnsig = sheet_obj1.cell(row = i, column = index_clinvar_clnsig)
       print(cell_obj1_clinvar_clnsig.value)

       cell_obj1_clinvar_review = sheet_obj1.cell(row = i, column = index_clinvar_review)
       print(cell_obj1_clinvar_review.value)

       cell_obj1_clinvar_pathogenicity = sheet_obj1.cell(row = i, column = index_clinvar_pathogenicity)
       print(cell_obj1_clinvar_pathogenicity.value)

       if (cell_obj1_clinvar_clnsig.value != "?") and (cell_obj1_clinvar_clnsig.value != "not_provided") and (cell_obj1_clinvar_clnsig.value != None):

          if ('Pathogenic' in cell_obj1_clinvar_clnsig.value or 'Likely_pathogenic' in cell_obj1_clinvar_clnsig.value) and (cell_obj1_clinvar_review.value != "criteria_provided,_single_submitter") and (cell_obj1_clinvar_review.value != "criteria_provided,_conflicting_interpretations") and (cell_obj1_clinvar_pathogenicity.value < 0.85):
             print(sheet_obj1.cell(row = i, column = 1).value, 'PS1: 1')
             sheet_obj1.cell(row = i, column = max_col1 + 1).value = 1
          else:
             print(sheet_obj1.cell(row = i, column = 1).value, 'PS1: 0')
             sheet_obj1.cell(row = i, column = max_col1 + 1).value = 0

       else:
             print(sheet_obj1.cell(row = i, column = 1).value, 'PS1: 0')
             sheet_obj1.cell(row = i, column = max_col1 + 1).value = 0


wb_obj1.save(output_vep_final_pathscore_workbook)



df_output_vep_final_pathscore = pd.read_excel(output_vep_final_pathscore_workbook)

df_5_newm = df_output_vep_final_pathscore.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PS1"], keep='first')

df_5_newn = df_5_newm.sort_values('PS1',ascending=True).drop_duplicates(subset=["#Uploaded_variation"], keep='last')
df_5_new = df_5_newn.sort_values('Location',ascending=True)

output_vep_final_pathscore_new_workbook = 'Enter your path\\output_vep_final_missense_path_new.xlsx'

df_5_new.to_excel(output_vep_final_pathscore_new_workbook, index=False)







output_vep_final_pathscore1_workbook = 'Enter your path\\output_vep_final_missense_path1.xlsx'


df_initial[['Amino_acids_start','Amino_acids_end']] = df_initial['Amino_acids'].str.split("/",expand=True)

df_initial.to_excel("Enter your path\\VEP_variants_Final.xlsx", index=False)

df_output_vep_clinvar_score[['Amino_acids_start','Amino_acids_end']] = df_output_vep_clinvar_score['Amino_acids'].str.split("/",expand=True)

df_output_vep_clinvar_score.to_excel(output_vep_clinvar_score_workbook, index=False)



df_6 = pd.merge(df_initial, df_output_vep_clinvar_score[["Consequence", "CDS_position", "Protein_position", "GIVEN_REF", "CLNREVSTAT", "CLNSIG", "CLNVI", "pathogenicity", "Amino_acids", "Amino_acids_start", "Amino_acids_end"]], on=['Consequence', 'CDS_position', 'Protein_position', 'GIVEN_REF', 'Amino_acids_start'], how='left')
#df_6 = pd.merge(df_initial, df_output_vep_clinvar_score[["Consequence", "CDS_position", "Protein_position", "HGVSc_ref", "CLNREVSTAT", "CLNSIG", "CLNVI", "pathogenicity", "Amino_acids", "Amino_acids_start", "Amino_acids_end"]], on=['Consequence', 'CDS_position', 'Protein_position', 'HGVSc_ref', 'Amino_acids_start'], how='left')


df_6.to_excel(output_vep_final_pathscore1_workbook, index=False)

print('The "output_vep_final_missense_path1.xlsx" file was saved to your path!')

wb_obj2 = openpyxl.load_workbook(output_vep_final_pathscore1_workbook)
   
sheet_obj2 = wb_obj2.active 

m_row2 = sheet_obj2.max_row
max_col2 = sheet_obj2.max_column


val19 = input("Enter the letter of column 'CLNSIG' in 'output_vep_final_missense_path1.xlsx' file: ") 
print(val19)
index_clinvar_clnsig = column_index_from_string(val19)


val20 = input("Enter the letter of column 'CLNREVSTAT' in 'output_vep_final_missense_path1.xlsx' file: ") 
print(val20)
index_clinvar_review = column_index_from_string(val20)

val21 = input("Enter the letter of column 'Amino_acids_end_x' in 'output_vep_final_missense_path1.xlsx' file: ") 
print(val21)
index_clinvar_amino_acids_end = column_index_from_string(val21)

val22 = input("Enter the letter of column 'Amino_acids_end_y' in 'output_vep_final_missense_path1.xlsx' file: ") 
print(val22)
index_clinvar_amino_acids_end1 = column_index_from_string(val22)

result_pm5 = sheet_obj2.cell(row= 1, column= max_col2 + 1, value='PM5')


for i in range(2, m_row2 + 1):

       cell_obj2_clinvar_clnsig = sheet_obj2.cell(row = i, column = index_clinvar_clnsig)
       print(cell_obj2_clinvar_clnsig.value)

       cell_obj2_clinvar_review = sheet_obj2.cell(row = i, column = index_clinvar_review)
       print(cell_obj2_clinvar_review.value)

       cell_obj2_clinvar_amino_acids_end = sheet_obj2.cell(row = i, column = index_clinvar_amino_acids_end)
       print(cell_obj2_clinvar_amino_acids_end.value)

       cell_obj2_clinvar_amino_acids_end1 = sheet_obj2.cell(row = i, column = index_clinvar_amino_acids_end1)
       print(cell_obj2_clinvar_amino_acids_end1.value)

       if (cell_obj2_clinvar_clnsig.value != "?") and (cell_obj2_clinvar_clnsig.value != "not_provided") and (cell_obj2_clinvar_clnsig.value != None):

          if ('Pathogenic' in cell_obj2_clinvar_clnsig.value or 'Likely_pathogenic' in cell_obj2_clinvar_clnsig.value) and (cell_obj2_clinvar_review.value != "criteria_provided,_single_submitter") and (cell_obj2_clinvar_review.value != "criteria_provided,_conflicting_interpretations") and (cell_obj2_clinvar_amino_acids_end.value != cell_obj2_clinvar_amino_acids_end1.value):
             print(sheet_obj2.cell(row = i, column = 1).value, 'PM5: 1')
             sheet_obj2.cell(row = i, column = max_col2 + 1).value = 1

          else:
             print(sheet_obj2.cell(row = i, column = 1).value, 'PM5: 0')
             sheet_obj2.cell(row = i, column = max_col2 + 1).value = 0

       else:
             print(sheet_obj2.cell(row = i, column = 1).value, 'PM5: 0')
             sheet_obj2.cell(row = i, column = max_col2 + 1).value = 0

wb_obj2.save(output_vep_final_pathscore1_workbook)



df_output_vep_final_pathscore1 = pd.read_excel(output_vep_final_pathscore1_workbook)


df_6_newm = df_output_vep_final_pathscore1.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PM5"], keep='first')

df_6_newn = df_6_newm.sort_values('PM5',ascending=True).drop_duplicates(subset=["#Uploaded_variation"], keep='last')
df_6_new = df_6_newn.sort_values('Location',ascending=True)

output_vep_final_pathscore1_new_workbook = 'Enter your path\\output_vep_final_missense_path1_new.xlsx'

df_6_new.to_excel(output_vep_final_pathscore1_new_workbook, index=False)





df_output_vep_final_pathscore_new = pd.read_excel(output_vep_final_pathscore_new_workbook)

df_output_vep_final_pathscore1_new = pd.read_excel(output_vep_final_pathscore1_new_workbook)




user_output1_workbook = input("Enter the path where the final annotation file and final features-scores will be saved: ")

input_file1 = output_annotation1.xlsx
output1_workbook = "{}\\{}".format(user_output1_workbook,input_file1)


df_initial = pd.read_excel(initial_workbook)

df_output_workbook = pd.read_excel(output_workbook)

df_ann1 = pd.merge(df_initial, df_output_workbook[["#Uploaded_variation", "POS", "REF", "ALT", "CLNDN", "CLNHGVS", "CLNREVSTAT", "CLNSIG", "CLNVI", "PP5", "BP6", "Repeat", "PM4", "BP3"]], on='#Uploaded_variation', how='left')

df_ann1.to_excel(output1_workbook, index=False)





input_file2 = output_annotation2.xlsx
output2_workbook = "{}\\{}".format(user_output1_workbook,input_file2)

df_Fisher_data = pd.read_excel(Fisher_exact_test_workbook)

df_ann2 = pd.merge(df_ann1, df_Fisher_data[["VariantOnGenome/DNA", "Case_Allele_Count", "case_total_all", "PS4", "FishersExact", "FishersExact_oddsratio", "FishersExact_pvalue"]], on='VariantOnGenome/DNA', how='outer')


df_ann2.to_excel(output2_workbook, index=False)



input_file3 = output_annotation3.xlsx
output3_workbook = "{}\\{}".format(user_output1_workbook,input_file3)

df_output_vep_final_pathscore_new = pd.read_excel(output_vep_final_pathscore_new_workbook)

df_ann3 = pd.merge(df_ann2, df_output_vep_final_pathscore_new[["#Uploaded_variation", "pathogenicity", "PS1"]], on='#Uploaded_variation', how='left')


df_ann3.to_excel(output3_workbook, index=False)



input_file4 = final_annotation_file.xlsx
final_annotation_workbook = "{}\\{}".format(user_output1_workbook,input_file4)

df_output_vep_final_pathscore1_new = pd.read_excel(output_vep_final_pathscore1_new_workbook)


df_ann4 = pd.merge(df_ann3, df_output_vep_final_pathscore1_new[["#Uploaded_variation", "Amino_acids_y", "Amino_acids_end_y", "PM5"]], on='#Uploaded_variation', how='left')


df_ann4.to_excel(final_annotation_workbook, index=False)



input_file_score = Final_scores.xlsx
final_score_workbook = "{}\\{}".format(user_output1_workbook,input_file_score)

input_file_score1 = Final_scores1.xlsx
final_score1_workbook = "{}\\{}".format(user_output1_workbook,input_file_score1)

input_file_score2 = Final_scores2.xlsx
final_score2_workbook = "{}\\{}".format(user_output1_workbook,input_file_score2)

input_file_score3 = Final_scores3.xlsx
final_score3_workbook = "{}\\{}".format(user_output1_workbook,input_file_score3)

input_file_score4 = Final_scores4.xlsx
final_score4_workbook = "{}\\{}".format(user_output1_workbook,input_file_score4)



#final_score_workbook = 'Enter your path\\Final_scores.xlsx'

#final_score1_workbook = 'Enter your path\\Final_scores1.xlsx'

#final_score2_workbook = 'Enter your path\\Final_scores2.xlsx'

#final_score3_workbook = 'Enter your path\\Final_scores3.xlsx'

#final_score4_workbook = 'Enter your path\\Final_scores4.xlsx'



df_final_score1 = pd.merge(df_initial[['#Uploaded_variation', 'Location', 'PVS1', 'BA1', 'BS1', 'PM2', 'PP3', 'BP4', 'BP7', 'PP2', 'BP1', 'PM1']], df_output_workbook[['#Uploaded_variation', 'PP5', 'BP6', 'PM4', 'BP3']], on='#Uploaded_variation', how='left')

df_final_score1_newn = df_final_score1.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3"], keep='first')
df_final_score1_new = df_final_score1_newn.sort_values('Location',ascending=True)

df_final_score1_new.to_excel(final_score1_workbook, index=False)




df_final_score2 = pd.merge(df_final_score1_new, df_Fisher_data[['#Uploaded_variation', 'PS4']], on='#Uploaded_variation', how='left')

df_final_score2_new = df_final_score2.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3", "PS4"], keep='first')

df_final_score2_new.to_excel(final_score2_workbook, index=False)



df_final_score3 = pd.merge(df_final_score2_new, df_output_vep_final_pathscore_new[['#Uploaded_variation', 'PS1']], on='#Uploaded_variation', how='left')
#df_final_score3 = pd.merge(df_final_score1_new, df_output_vep_final_pathscore_new[['#Uploaded_variation', 'PS1']], on='#Uploaded_variation', how='left')


df_final_score3_newn = df_final_score3.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3", "PS4", "PS1"], keep='first')

df_final_score3_new = df_final_score3_newn.sort_values('Location',ascending=True)

df_final_score3_new.to_excel(final_score3_workbook, index=False)



df_final_score4 = pd.merge(df_final_score3_new, df_output_vep_final_pathscore1_new[['#Uploaded_variation', 'PM5']], on='#Uploaded_variation', how='left')

df_final_score4_newn = df_final_score4.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3", "PS4", "PS1", "PM5"], keep='first')

df_final_score4_new = df_final_score4_newn.sort_values('Location',ascending=True)

df_final_score4_new.to_excel(final_score4_workbook, index=False)





#df_final_score_newn = df_final_score.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3", "PS4", "PS1", "PM5", "PS2", "PM6", "PS3", "BS3", "PM3", "BP2", "PP1", "BS4", "PP4", "BP5", "BS2"], keep='first')
##df_final_score_newn = df_final_score.sort_values('#Uploaded_variation',ascending=True).drop_duplicates(subset=["#Uploaded_variation", "PVS1", "BA1", "BS1", "PM2", "PP3", "BP4", "BP7", "PP2", "BP1", "PM1", "PP5", "BP6", "PM4", "BP3", "PS1", "PM5", "PS2", "PM6", "PS3", "BS3", "PM3", "BP2", "PP1", "BS4", "PP4", "BP5", "BS2"], keep='first')
#df_final_score_new = df_final_score_newn.sort_values('Location',ascending=True)

#df_final_score_new.to_excel(final_score_workbook, index=False)



df_final_score4_new[['chromosome','VariantOnGenome/DNA']] = df_final_score4_new['#Uploaded_variation'].str.split(":",expand=True)
#df_final_score_new[['id_ncbi','VariantOnTranscript/DNA']] = df_final_score_new['#Uploaded_variation'].str.split(":",expand=True)

df_final_score4_new.to_excel("Enter your path\\Final_scores.xlsx", index=False)



labels_workbook = 'Enter your path\\samples_labels.xlsx'

df_labels = pd.read_excel(labels_workbook)
df_labels = df_labels.sort_values('VariantOnGenome/DNA',ascending=True)
df_labels.to_excel(labels_workbook, index=False)

final_score_labels_workbook = 'Enter your path\\Final_features_labels.xlsx'


df_final_score_labels = pd.merge(df_final_score4_new[['VariantOnGenome/DNA', 'PVS1', 'BA1', 'BS1', 'PM2', 'PP3', 'BP4', 'BP7', 'PP2', 'BP1', 'PM1', 'PP5', 'BP6', 'PM4', 'BP3', 'PS4', 'PS1', 'PM5']], df_labels[['VariantOnGenome/DNA', 'effectid']], on='VariantOnGenome/DNA', how='left')

#df_final_score_labels = pd.merge(df_final_score4_new[['VariantOnGenome/DNA', 'PVS1', 'BA1', 'BS1', 'PM2', 'PP3', 'BP4', 'BP7', 'PP2', 'BP1', 'PM1', 'PP5', 'BP6', 'PM4', 'BP3', 'PS4', 'PS1', 'PM5', 'PS2', 'PM6', 'PS3', 'BS3', 'PM3', 'BP2', 'PP1', 'BS4', 'PP4', 'BP5', 'BS2']], df_labels[['VariantOnGenome/DNA', 'effectid']], on='VariantOnGenome/DNA', how='left')


df_final_score_labels.to_excel(final_score_labels_workbook, index=False)






